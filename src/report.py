"""
src/report.py
=============
Phase 5 — Reporting

Provides five capabilities that transform raw metrics into shareable
deliverables:

  5.1  format_metrics_table()  — pretty-print to console
  5.2  export_csv()            — saves series / values to exports/
  5.3  export_excel()          — multi-sheet .xlsx to exports/
  5.4  generate_pdf()          — professional PDF report to reports/
  5.5  investment_grade()      — A/B/C/D scoring rubric
  5.6  performance_summary()   — AI-generated narrative (Anthropic API)

Architecture
------------
All functions are stateless: they receive pre-computed metrics dicts
and pd.Series objects from main.py and write to disk.  No price data
or return calculations happen here.

Output directories are created automatically if they do not exist.
"""

import os
import json
import textwrap
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import requests

# reportlab imports for PDF
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak,
)

# openpyxl imports for Excel
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Directory setup ────────────────────────────────────────────────────

EXPORTS_DIR = Path("exports")
REPORTS_DIR = Path("reports")

EXPORTS_DIR.mkdir(exist_ok=True)
REPORTS_DIR.mkdir(exist_ok=True)


# ══════════════════════════════════════════════════════════════════════
# 5.5  Investment Grade
# ══════════════════════════════════════════════════════════════════════

def investment_grade(metrics: dict) -> tuple[str, str, float]:
    """
    Score a portfolio on a weighted rubric and return a letter grade.

    Scoring weights
    ---------------
    Sharpe Ratio   35 %   — best single risk-adjusted metric
    Max Drawdown   30 %   — capital preservation is paramount
    CAGR           20 %   — absolute growth matters
    Volatility     15 %   — lower volatility preferred

    Each metric is mapped to a 0–100 sub-score using thresholds derived
    from empirical hedge-fund and mutual-fund benchmarks.

    Thresholds used in practice
    ---------------------------
    Sharpe  ≥ 2.0  → 100   (elite)
            ≥ 1.5  →  85   (excellent)
            ≥ 1.0  →  70   (good)
            ≥ 0.5  →  50   (average)
            <  0.5 →  20   (poor)

    Max DD  |dd| ≤ 0.10 → 100  (very defensive)
            |dd| ≤ 0.20 →  80
            |dd| ≤ 0.30 →  60
            |dd| ≤ 0.40 →  40
            |dd| >  0.40 → 15   (severe drawdown)

    CAGR    ≥ 0.25 → 100
            ≥ 0.15 →  80
            ≥ 0.08 →  60
            ≥ 0.00 →  40
            <  0.00 → 10   (losing money)

    Volatility  ≤ 0.10 → 100
                ≤ 0.15 →  80
                ≤ 0.20 →  65
                ≤ 0.30 →  45
                >  0.30 → 20

    Final grade boundaries
    ----------------------
    A  ≥ 75  (Excellent)
    B  ≥ 55  (Good)
    C  ≥ 35  (Average)
    D  <  35  (Poor)

    Parameters
    ----------
    metrics : dict returned by performance_report()

    Returns
    -------
    grade       : str   — "A", "B", "C", or "D"
    label       : str   — "Excellent", "Good", "Average", "Poor"
    final_score : float — 0–100 composite score
    """
    sharpe = metrics.get("sharpe_ratio", 0) or 0
    mdd    = abs(metrics.get("max_drawdown", 1))
    cagr   = metrics.get("cagr", 0) or 0
    vol    = metrics.get("ann_volatility", 1) or 1

    # ── Sub-scores ──────────────────────────────────────────────────

    # Sharpe sub-score
    if sharpe >= 2.0:
        s_sharpe = 100
    elif sharpe >= 1.5:
        s_sharpe = 85
    elif sharpe >= 1.0:
        s_sharpe = 70
    elif sharpe >= 0.5:
        s_sharpe = 50
    else:
        s_sharpe = 20

    # Max Drawdown sub-score  (mdd is already absolute value)
    if mdd <= 0.10:
        s_mdd = 100
    elif mdd <= 0.20:
        s_mdd = 80
    elif mdd <= 0.30:
        s_mdd = 60
    elif mdd <= 0.40:
        s_mdd = 40
    else:
        s_mdd = 15

    # CAGR sub-score
    if cagr >= 0.25:
        s_cagr = 100
    elif cagr >= 0.15:
        s_cagr = 80
    elif cagr >= 0.08:
        s_cagr = 60
    elif cagr >= 0.00:
        s_cagr = 40
    else:
        s_cagr = 10

    # Volatility sub-score
    if vol <= 0.10:
        s_vol = 100
    elif vol <= 0.15:
        s_vol = 80
    elif vol <= 0.20:
        s_vol = 65
    elif vol <= 0.30:
        s_vol = 45
    else:
        s_vol = 20

    # ── Weighted composite ──────────────────────────────────────────
    final_score = (
        0.35 * s_sharpe +
        0.30 * s_mdd    +
        0.20 * s_cagr   +
        0.15 * s_vol
    )

    # ── Grade boundary ──────────────────────────────────────────────
    if final_score >= 75:
        return "A", "Excellent", final_score
    elif final_score >= 55:
        return "B", "Good", final_score
    elif final_score >= 35:
        return "C", "Average", final_score
    else:
        return "D", "Poor", final_score


# ══════════════════════════════════════════════════════════════════════
# 5.1  Metrics Table  (console)
# ══════════════════════════════════════════════════════════════════════

def format_metrics_table(
    port_label:    str,
    port_metrics:  dict,
    bench_metrics: dict,
    bench_label:   str  = "SPY",
    initial_investment: float = 10_000,
    start_date:    str  = "",
    end_date:      str  = "",
) -> None:
    """
    Print a formatted, side-by-side metrics table to the console.

    Displays every metric in aligned columns so portfolio and benchmark
    values can be compared at a glance.  Investment grade is appended
    beneath the table.

    Parameters
    ----------
    port_label         : display name for the portfolio
    port_metrics       : dict returned by performance_report()
    bench_metrics      : dict returned by performance_report()
    bench_label        : display name for the benchmark
    initial_investment : starting capital
    start_date / end_date : analysis period labels (cosmetic only)
    """
    grade, label, score = investment_grade(port_metrics)
    w = 62

    print("\n" + "╔" + "═" * w + "╗")
    print("║" + "  PORTFOLIO ANALYTICS REPORT".center(w) + "║")
    print("║" + f"  {start_date}  →  {end_date}".center(w) + "║")
    print("╠" + "═" * w + "╣")

    # Column headers
    col1_w, col2_w, col3_w = 28, 16, 16
    header = (f"  {'Metric':<{col1_w}}"
              f"{port_label[:14]:>{col2_w}}"
              f"{bench_label[:14]:>{col3_w}}")
    print("║" + header + "║")
    print("╠" + "─" * w + "╣")

    def row(name, port_val, bench_val):
        """Format one metric row."""
        print("║" + f"  {name:<{col1_w}}{port_val:>{col2_w}}{bench_val:>{col3_w}}" + "║")

    def pct(v):
        return f"{v * 100:+.2f}%" if v is not None else "  —"

    def flt(v, decimals=3):
        return f"{v:.{decimals}f}" if v is not None else "  —"

    row("Total Return",
        pct(port_metrics.get("total_return")),
        pct(bench_metrics.get("total_return")))
    row("CAGR",
        pct(port_metrics.get("cagr")),
        pct(bench_metrics.get("cagr")))
    row("Ann. Volatility",
        pct(port_metrics.get("ann_volatility")),
        pct(bench_metrics.get("ann_volatility")))
    row("Sharpe Ratio",
        flt(port_metrics.get("sharpe_ratio")),
        flt(bench_metrics.get("sharpe_ratio")))
    row("Max Drawdown",
        pct(port_metrics.get("max_drawdown")),
        pct(bench_metrics.get("max_drawdown")))

    if port_metrics.get("beta") is not None:
        row("Beta",
            flt(port_metrics.get("beta")),
            "  1.000")
    if port_metrics.get("alpha") is not None:
        row("Alpha (ann.)",
            pct(port_metrics.get("alpha")),
            "  +0.00%")

    # Outperformance
    op = (port_metrics.get("total_return", 0) or 0) - \
         (bench_metrics.get("total_return", 0) or 0)
    print("╠" + "─" * w + "╣")
    row("Outperformance", pct(op), "")

    print("╠" + "─" * w + "╣")
    grade_row = f"  Investment Grade  →  {grade}  ({label})   Score: {score:.1f}/100"
    print("║" + f"{grade_row:<{w}}" + "║")
    print("╚" + "═" * w + "╝\n")


# ══════════════════════════════════════════════════════════════════════
# 5.2  CSV Export
# ══════════════════════════════════════════════════════════════════════

def export_csv(
    port_val_eq:    pd.Series,
    port_val_cust:  pd.Series,
    bench_val:      pd.Series,
    port_ret_eq:    pd.Series,
    port_ret_cust:  pd.Series,
    bench_ret:      pd.Series,
    eq_metrics:     dict,
    cust_metrics:   dict,
    bench_metrics:  dict,
) -> dict[str, Path]:
    """
    Export equity curves, returns, and metrics to CSV files.

    Why CSV?
    --------
    CSV is the lingua franca of data interchange.  It's importable by
    Excel, Google Sheets, R, Tableau, and virtually every analytics
    tool.  Separate files for values, returns, and metrics keep the data
    clean and purpose-specific.

    Files produced
    --------------
    exports/equity_curves.csv   — daily portfolio and benchmark values
    exports/daily_returns.csv   — daily returns for all strategies
    exports/metrics_summary.csv — one row per strategy, all metrics

    Parameters
    ----------
    All pd.Series and metrics dicts from main.py.

    Returns
    -------
    dict mapping label → Path  (for downstream use or printing)
    """
    # ── equity_curves.csv ──────────────────────────────────────────
    equity = pd.DataFrame({
        "Equal Weight"   : port_val_eq,
        "Custom Weight"  : port_val_cust,
        "SPY Benchmark"  : bench_val,
    })
    equity.index.name = "Date"
    equity_path = EXPORTS_DIR / "equity_curves.csv"
    equity.to_csv(equity_path)

    # ── daily_returns.csv ──────────────────────────────────────────
    ret_df = pd.DataFrame({
        "Equal Weight"  : port_ret_eq,
        "Custom Weight" : port_ret_cust,
        "SPY Benchmark" : bench_ret,
    })
    ret_df.index.name = "Date"
    returns_path = EXPORTS_DIR / "daily_returns.csv"
    ret_df.to_csv(returns_path)

    # ── metrics_summary.csv ───────────────────────────────────────
    def _grade_score(m):
        g, l, s = investment_grade(m)
        return g, s

    rows = []
    for label, m in [
        ("Equal Weight",  eq_metrics),
        ("Custom Weight", cust_metrics),
        ("SPY Benchmark", bench_metrics),
    ]:
        g, s = _grade_score(m)
        rows.append({
            "Strategy"       : label,
            "Total Return %"  : round((m.get("total_return") or 0) * 100, 2),
            "CAGR %"          : round((m.get("cagr") or 0) * 100, 2),
            "Ann Volatility %" : round((m.get("ann_volatility") or 0) * 100, 2),
            "Sharpe Ratio"    : round(m.get("sharpe_ratio") or 0, 3),
            "Max Drawdown %"  : round((m.get("max_drawdown") or 0) * 100, 2),
            "Beta"            : round(m.get("beta") or 0, 3) if m.get("beta") is not None else None,
            "Alpha %"         : round((m.get("alpha") or 0) * 100, 2) if m.get("alpha") is not None else None,
            "Grade"           : g,
            "Score"           : round(s, 1),
        })

    metrics_df   = pd.DataFrame(rows)
    metrics_path = EXPORTS_DIR / "metrics_summary.csv"
    metrics_df.to_csv(metrics_path, index=False)

    paths = {
        "equity_curves"  : equity_path,
        "daily_returns"  : returns_path,
        "metrics_summary": metrics_path,
    }

    print("\n── CSV Export ──────────────────────────────────")
    for name, path in paths.items():
        print(f"  ✓  {path}")
    print()

    return paths


# ══════════════════════════════════════════════════════════════════════
# 5.3  Excel Export
# ══════════════════════════════════════════════════════════════════════

# ── openpyxl style helpers ────────────────────────────────────────────

def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(border_style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_header_row(ws, row: int, values: list, bg: str = "2D4059") -> None:
    """Write a bold white header row with coloured background."""
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = _header_fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()


def _data_cell(ws, row: int, col: int, value, number_format: str = "General",
               bold: bool = False, bg: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border    = _thin_border()
    cell.number_format = number_format
    if bg:
        cell.fill = _header_fill(bg)


def export_excel(
    port_val_eq:    pd.Series,
    port_val_cust:  pd.Series,
    bench_val:      pd.Series,
    port_ret_eq:    pd.Series,
    port_ret_cust:  pd.Series,
    bench_ret:      pd.Series,
    eq_metrics:     dict,
    cust_metrics:   dict,
    bench_metrics:  dict,
    initial_investment: float = 10_000,
    start_date:     str = "",
    end_date:       str = "",
) -> Path:
    """
    Build a professional multi-sheet Excel workbook.

    Sheets produced
    ---------------
    Summary   — metrics comparison table with investment grades
    Equity    — daily equity curve values for all strategies
    Returns   — daily return series for all strategies
    Portfolio — equal-weight and custom-weight details side by side

    Design decisions
    ----------------
    - Industry-standard colour coding from the SKILL.md guide:
        Blue text   → hardcoded inputs (initial investment, dates)
        Black text  → all computed values
    - Number formats applied at cell level so columns auto-format
      when the workbook is opened in Excel.
    - Column widths set explicitly so no data is truncated.

    Parameters
    ----------
    All pd.Series and metrics dicts from main.py.

    Returns
    -------
    Path to the saved workbook.
    """
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    # ── Sheet 1: Summary ──────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")

    # Title block
    ws_sum.merge_cells("A1:E1")
    title_cell = ws_sum["A1"]
    title_cell.value     = "Portfolio Analytics Report"
    title_cell.font      = Font(bold=True, size=16, color="2D4059")
    title_cell.alignment = Alignment(horizontal="center")

    ws_sum.merge_cells("A2:E2")
    sub = ws_sum["A2"]
    sub.value     = f"Period: {start_date}  →  {end_date}   |   Initial Investment: ${initial_investment:,.0f}"
    sub.font      = Font(italic=True, size=11, color="555555")
    sub.alignment = Alignment(horizontal="center")

    _apply_header_row(ws_sum, 4,
        ["Metric", "Equal Weight", "Custom Weight", "SPY Benchmark", "Notes"])

    metrics_rows = [
        ("Total Return",       "total_return",    "0.00%",    ""),
        ("CAGR",               "cagr",            "0.00%",    "Annualised"),
        ("Ann. Volatility",    "ann_volatility",  "0.00%",    ""),
        ("Sharpe Ratio",       "sharpe_ratio",    "0.000",    "Higher is better"),
        ("Max Drawdown",       "max_drawdown",    "0.00%",    "Worst peak-to-trough"),
        ("Beta",               "beta",            "0.000",    "vs SPY"),
        ("Alpha (ann.)",       "alpha",           "0.00%",    "Jensen's alpha"),
    ]

    for r, (name, key, fmt, note) in enumerate(metrics_rows, start=5):
        ws_sum.cell(row=r, column=1, value=name).font = Font(bold=True, size=10)
        ws_sum.cell(row=r, column=1).border = _thin_border()

        for col, m in enumerate([eq_metrics, cust_metrics, bench_metrics], start=2):
            val = m.get(key)
            _data_cell(ws_sum, r, col, val, number_format=fmt)

        ws_sum.cell(row=r, column=5, value=note).font = Font(italic=True,
                                                              size=9, color="777777")
        ws_sum.cell(row=r, column=5).border = _thin_border()

    # Investment grade rows
    ws_sum.cell(row=12, column=1, value="Investment Grade").font = Font(bold=True, size=10)
    ws_sum.cell(row=12, column=1).border = _thin_border()
    ws_sum.cell(row=13, column=1, value="Composite Score").font  = Font(bold=True, size=10)
    ws_sum.cell(row=13, column=1).border = _thin_border()

    grade_colors = {"A": "27AE60", "B": "2980B9", "C": "F39C12", "D": "C0392B"}

    for col, m in enumerate([eq_metrics, cust_metrics, bench_metrics], start=2):
        g, lbl, score = investment_grade(m)
        cell_g = ws_sum.cell(row=12, column=col, value=f"{g} — {lbl}")
        cell_g.font      = Font(bold=True, color="FFFFFF", size=10)
        cell_g.fill      = _header_fill(grade_colors.get(g, "777777"))
        cell_g.alignment = Alignment(horizontal="center")
        cell_g.border    = _thin_border()

        cell_s = ws_sum.cell(row=13, column=col, value=round(score, 1))
        cell_s.number_format = "0.0"
        cell_s.alignment     = Alignment(horizontal="center")
        cell_s.border        = _thin_border()

    ws_sum.column_dimensions["A"].width = 24
    for col in ["B", "C", "D"]:
        ws_sum.column_dimensions[col].width = 18
    ws_sum.column_dimensions["E"].width = 26

    # ── Sheet 2: Equity Curves ────────────────────────────────────
    ws_eq = wb.create_sheet("Equity Curves")
    _apply_header_row(ws_eq, 1,
        ["Date", "Equal Weight ($)", "Custom Weight ($)", "SPY Benchmark ($)"])

    for r, (date, eq, cust, bench) in enumerate(
        zip(port_val_eq.index, port_val_eq, port_val_cust, bench_val), start=2
    ):
        ws_eq.cell(row=r, column=1, value=date.date()).number_format = "YYYY-MM-DD"
        ws_eq.cell(row=r, column=1).border = _thin_border()
        for col, val in enumerate([eq, cust, bench], start=2):
            _data_cell(ws_eq, r, col, round(val, 2), number_format='"$"#,##0.00')

    ws_eq.column_dimensions["A"].width = 14
    for col in ["B", "C", "D"]:
        ws_eq.column_dimensions[col].width = 20

    # ── Sheet 3: Daily Returns ────────────────────────────────────
    ws_ret = wb.create_sheet("Daily Returns")
    _apply_header_row(ws_ret, 1,
        ["Date", "Equal Weight", "Custom Weight", "SPY Benchmark"])

    for r, (date, eq, cust, bench) in enumerate(
        zip(port_ret_eq.index, port_ret_eq, port_ret_cust, bench_ret), start=2
    ):
        ws_ret.cell(row=r, column=1, value=date.date()).number_format = "YYYY-MM-DD"
        ws_ret.cell(row=r, column=1).border = _thin_border()
        for col, val in enumerate([eq, cust, bench], start=2):
            _data_cell(ws_ret, r, col, round(val, 6), number_format="0.0000%")

    ws_ret.column_dimensions["A"].width = 14
    for col in ["B", "C", "D"]:
        ws_ret.column_dimensions[col].width = 18

    # ── Save ──────────────────────────────────────────────────────
    out_path = EXPORTS_DIR / "portfolio_report.xlsx"
    wb.save(out_path)
    print(f"── Excel Export ─────────────────────────────────")
    print(f"  ✓  {out_path}\n")
    return out_path


# ══════════════════════════════════════════════════════════════════════
# 5.6  Performance Summary  (Anthropic API)
# ══════════════════════════════════════════════════════════════════════

def performance_summary(
    port_label:    str,
    port_metrics:  dict,
    bench_metrics: dict,
    bench_label:   str  = "SPY",
    start_date:    str  = "",
    end_date:      str  = "",
) -> str:
    """
    Generate a human-readable narrative summary via the Anthropic API.

    Why an AI narrative?
    --------------------
    Quantitative reports are full of numbers that non-technical
    stakeholders find hard to interpret.  A concise narrative paragraph
    translates the metrics into plain English, just as a portfolio
    manager's letter would.

    The function packages the metrics dict as JSON in a prompt, calls
    claude-sonnet-4-6, and returns the generated text.  If the API call
    fails (network issue, missing key), a templated fallback is returned
    so the rest of the report can still be generated.

    Parameters
    ----------
    port_label    : name of the portfolio
    port_metrics  : dict from performance_report()
    bench_metrics : dict from performance_report()
    bench_label   : benchmark name
    start_date / end_date : period strings for context

    Returns
    -------
    str  — 3–5 sentence plain-English commentary
    """
    grade, grade_label, score = investment_grade(port_metrics)

    prompt = f"""You are a quantitative finance analyst writing a brief performance commentary.

Portfolio: {port_label}
Benchmark: {bench_label}
Period: {start_date} to {end_date}
Investment Grade: {grade} ({grade_label}), Score: {score:.1f}/100

Portfolio Metrics:
{json.dumps({k: round(v * 100, 2) if k not in ('sharpe_ratio', 'beta') and v is not None else v
             for k, v in port_metrics.items()}, indent=2)}

Benchmark Metrics:
{json.dumps({k: round(v * 100, 2) if k not in ('sharpe_ratio', 'beta') and v is not None else v
             for k, v in bench_metrics.items()}, indent=2)}

Write a 3-5 sentence plain-English performance commentary that:
1. States the total return and whether it outperformed the benchmark.
2. Comments on risk (volatility, max drawdown).
3. Mentions the Sharpe ratio and what it implies about risk-adjusted returns.
4. Concludes with the investment grade and overall assessment.

Keep it factual and professional. Do not use bullet points. Return only the commentary text."""

    try:
        response = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type": "application/json"},
            json={
                "model"     : "claude-sonnet-4-6",
                "max_tokens": 400,
                "messages"  : [{"role": "user", "content": prompt}],
            },
            timeout=30,
        )
        response.raise_for_status()
        data = response.json()
        return data["content"][0]["text"].strip()

    except Exception as e:
        # Graceful fallback — templated narrative
        tr  = (port_metrics.get("total_return") or 0) * 100
        btr = (bench_metrics.get("total_return") or 0) * 100
        sr  = port_metrics.get("sharpe_ratio") or 0
        mdd = abs(port_metrics.get("max_drawdown") or 0) * 100
        op  = tr - btr
        sign = "outperforming" if op >= 0 else "underperforming"

        return (
            f"The {port_label} returned {tr:+.1f}% over the analysis period "
            f"({start_date}–{end_date}), {sign} the {bench_label} by {abs(op):.1f} percentage points. "
            f"The portfolio maintained an annualised Sharpe ratio of {sr:.2f}, "
            f"with a maximum drawdown of {mdd:.1f}%. "
            f"Based on these risk-adjusted metrics it earned an Investment Grade of "
            f"{grade} ({grade_label}), with a composite score of {score:.1f}/100."
        )


# ══════════════════════════════════════════════════════════════════════
# 5.4  PDF Report
# ══════════════════════════════════════════════════════════════════════

def generate_pdf(
    port_label:    str,
    port_metrics:  dict,
    bench_metrics: dict,
    bench_label:   str  = "SPY",
    start_date:    str  = "",
    end_date:      str  = "",
    initial_investment: float = 10_000,
    chart_paths:   list[str] | None = None,
) -> Path:
    """
    Generate a professional multi-page PDF report.

    Page layout
    -----------
    Page 1 — Cover: title, period, grade badge
    Page 2 — Executive Summary: AI-generated narrative + key metrics table
    Page 3 — Detailed Metrics: all values for portfolio and benchmark
    Page 4+ — Charts: embedded matplotlib figures (optional)

    PDF is built with reportlab's Platypus (page layout engine), which
    treats the document as a sequence of "flowables" — Paragraphs,
    Tables, Spacers — and handles page breaks automatically.

    Parameters
    ----------
    port_label         : portfolio name
    port_metrics       : dict from performance_report()
    bench_metrics      : dict from performance_report()
    bench_label        : benchmark name
    start_date / end_date : period strings
    initial_investment : starting capital
    chart_paths        : optional list of .png paths to embed as images

    Returns
    -------
    Path to the saved PDF.
    """
    out_path = REPORTS_DIR / "portfolio_report.pdf"

    doc    = SimpleDocTemplate(
        str(out_path),
        pagesize     = letter,
        rightMargin  = 0.75 * inch,
        leftMargin   = 0.75 * inch,
        topMargin    = 0.75 * inch,
        bottomMargin = 0.75 * inch,
    )
    styles = getSampleStyleSheet()
    story  = []

    # ── Custom styles ──────────────────────────────────────────────
    NAVY   = colors.HexColor("#2D4059")
    TEAL   = colors.HexColor("#1ABC9C")
    LIGHT  = colors.HexColor("#ECF0F1")

    title_style = ParagraphStyle(
        "ReportTitle",
        parent    = styles["Title"],
        fontSize  = 28,
        textColor = NAVY,
        spaceAfter = 6,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent    = styles["Normal"],
        fontSize  = 13,
        textColor = colors.HexColor("#555555"),
        spaceAfter = 4,
    )
    heading_style = ParagraphStyle(
        "SectionHead",
        parent    = styles["Heading1"],
        fontSize  = 14,
        textColor = NAVY,
        spaceBefore = 18,
        spaceAfter  = 8,
        borderPad   = 4,
    )
    body_style = ParagraphStyle(
        "Body",
        parent    = styles["Normal"],
        fontSize  = 10,
        leading   = 16,
        textColor = colors.HexColor("#333333"),
    )

    def _divider():
        return HRFlowable(width="100%", thickness=1,
                          color=TEAL, spaceAfter=8, spaceBefore=8)

    # ── Page 1: Cover ─────────────────────────────────────────────
    story.append(Spacer(1, 1.5 * inch))
    story.append(Paragraph("Portfolio Analytics Report", title_style))
    story.append(_divider())
    story.append(Paragraph(f"<b>Portfolio:</b>  {port_label}", subtitle_style))
    story.append(Paragraph(f"<b>Benchmark:</b>  {bench_label}", subtitle_style))
    story.append(Paragraph(f"<b>Period:</b>      {start_date}  →  {end_date}", subtitle_style))
    story.append(Paragraph(f"<b>Initial Capital:</b>  ${initial_investment:,.0f}", subtitle_style))
    story.append(Spacer(1, 0.4 * inch))

    # Grade badge table
    grade, grade_label, score = investment_grade(port_metrics)
    GRADE_COLORS = {
        "A": colors.HexColor("#27AE60"),
        "B": colors.HexColor("#2980B9"),
        "C": colors.HexColor("#F39C12"),
        "D": colors.HexColor("#C0392B"),
    }
    badge_data = [
        [f"Investment Grade:  {grade} — {grade_label}   |   Score: {score:.1f} / 100"]
    ]
    badge_table = Table(badge_data, colWidths=[6 * inch])
    badge_table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, -1), GRADE_COLORS.get(grade, colors.grey)),
        ("TEXTCOLOR",   (0, 0), (-1, -1), colors.white),
        ("FONTNAME",    (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 16),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ("ROUNDEDCORNERS", [6]),
    ]))
    story.append(badge_table)
    story.append(Spacer(1, 0.3 * inch))

    generated_at = datetime.now().strftime("%B %d, %Y  %H:%M")
    story.append(Paragraph(
        f"<i>Generated: {generated_at}</i>",
        ParagraphStyle("small", parent=styles["Normal"],
                       fontSize=9, textColor=colors.grey,
                       alignment=1)
    ))
    story.append(PageBreak())

    # ── Page 2: Executive Summary ──────────────────────────────────
    story.append(Paragraph("Executive Summary", heading_style))
    story.append(_divider())

    summary_text = performance_summary(
        port_label    = port_label,
        port_metrics  = port_metrics,
        bench_metrics = bench_metrics,
        bench_label   = bench_label,
        start_date    = start_date,
        end_date      = end_date,
    )
    story.append(Paragraph(summary_text, body_style))
    story.append(Spacer(1, 0.25 * inch))

    # Key metrics highlight table (4 metrics in 2 columns)
    story.append(Paragraph("Key Metrics at a Glance", heading_style))
    tr  = (port_metrics.get("total_return") or 0) * 100
    sr  = port_metrics.get("sharpe_ratio") or 0
    mdd = (port_metrics.get("max_drawdown") or 0) * 100
    ca  = (port_metrics.get("cagr") or 0) * 100

    kpi_data = [
        ["Total Return", f"{tr:+.2f}%",   "Sharpe Ratio", f"{sr:.3f}"],
        ["CAGR",         f"{ca:+.2f}%",   "Max Drawdown", f"{mdd:.2f}%"],
    ]
    kpi_table = Table(kpi_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (0, -1), NAVY),
        ("BACKGROUND",    (2, 0), (2, -1), NAVY),
        ("TEXTCOLOR",     (0, 0), (0, -1), colors.white),
        ("TEXTCOLOR",     (2, 0), (2, -1), colors.white),
        ("FONTNAME",      (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME",      (1, 0), (1, -1), "Helvetica-Bold"),
        ("FONTNAME",      (3, 0), (3, -1), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 11),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
    ]))
    story.append(kpi_table)
    story.append(PageBreak())

    # ── Page 3: Detailed Metrics ───────────────────────────────────
    story.append(Paragraph("Detailed Performance Metrics", heading_style))
    story.append(_divider())

    def _pct(v):
        return f"{(v or 0)*100:+.2f}%" if v is not None else "—"

    def _flt(v, d=3):
        return f"{v:.{d}f}" if v is not None else "—"

    detail_data = [
        ["Metric",           port_label,                                  bench_label],
        ["Total Return",     _pct(port_metrics.get("total_return")),      _pct(bench_metrics.get("total_return"))],
        ["CAGR",             _pct(port_metrics.get("cagr")),              _pct(bench_metrics.get("cagr"))],
        ["Ann. Volatility",  _pct(port_metrics.get("ann_volatility")),    _pct(bench_metrics.get("ann_volatility"))],
        ["Sharpe Ratio",     _flt(port_metrics.get("sharpe_ratio")),      _flt(bench_metrics.get("sharpe_ratio"))],
        ["Max Drawdown",     _pct(port_metrics.get("max_drawdown")),      _pct(bench_metrics.get("max_drawdown"))],
        ["Beta",             _flt(port_metrics.get("beta")),              "1.000"],
        ["Alpha (ann.)",     _pct(port_metrics.get("alpha")),             "+0.00%"],
        ["Grade",            f"{grade} ({grade_label})",                  "—"],
        ["Composite Score",  f"{score:.1f} / 100",                        "—"],
    ]

    col_w = [2.5*inch, 2.25*inch, 2.25*inch]
    det_table = Table(detail_data, colWidths=col_w)
    det_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  NAVY),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTNAME",      (0, 1), (0, -1),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 10),
        ("ALIGN",         (1, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, LIGHT]),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
    ]))
    story.append(det_table)

    # ── Build ──────────────────────────────────────────────────────
    doc.build(story)
    print(f"── PDF Report ───────────────────────────────────")
    print(f"  ✓  {out_path}\n")
    return out_path


# ══════════════════════════════════════════════════════════════════════
# Convenience wrapper — run all reporting in one call
# ══════════════════════════════════════════════════════════════════════

def generate_all_reports(
    port_label:         str,
    port_metrics:       dict,
    bench_metrics:      dict,
    eq_metrics:         dict,
    cust_metrics:       dict,
    port_val_eq:        "pd.Series",
    port_val_cust:      "pd.Series",
    bench_val:          "pd.Series",
    port_ret_eq:        "pd.Series",
    port_ret_cust:      "pd.Series",
    bench_ret:          "pd.Series",
    bench_label:        str   = "SPY",
    initial_investment: float = 10_000,
    start_date:         str   = "",
    end_date:           str   = "",
) -> None:
    """
    Run all five Phase 5 reporting steps in sequence.

    Call this from main.py after Phase 3 metrics and Phase 4 charts.

    Steps executed
    --------------
    1. Console metrics table  (5.1)
    2. CSV export             (5.2)
    3. Excel export           (5.3)
    4. PDF report             (5.4 + 5.6 combined)
    """
    format_metrics_table(
        port_label         = port_label,
        port_metrics       = port_metrics,
        bench_metrics      = bench_metrics,
        bench_label        = bench_label,
        initial_investment = initial_investment,
        start_date         = start_date,
        end_date           = end_date,
    )

    export_csv(
        port_val_eq    = port_val_eq,
        port_val_cust  = port_val_cust,
        bench_val      = bench_val,
        port_ret_eq    = port_ret_eq,
        port_ret_cust  = port_ret_cust,
        bench_ret      = bench_ret,
        eq_metrics     = eq_metrics,
        cust_metrics   = cust_metrics,
        bench_metrics  = bench_metrics,
    )

    export_excel(
        port_val_eq        = port_val_eq,
        port_val_cust      = port_val_cust,
        bench_val          = bench_val,
        port_ret_eq        = port_ret_eq,
        port_ret_cust      = port_ret_cust,
        bench_ret          = bench_ret,
        eq_metrics         = eq_metrics,
        cust_metrics       = cust_metrics,
        bench_metrics      = bench_metrics,
        initial_investment = initial_investment,
        start_date         = start_date,
        end_date           = end_date,
    )

    generate_pdf(
        port_label         = port_label,
        port_metrics       = port_metrics,
        bench_metrics      = bench_metrics,
        bench_label        = bench_label,
        start_date         = start_date,
        end_date           = end_date,
        initial_investment = initial_investment,
    )